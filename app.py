import io
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed

import google.auth
import google.auth.transport.requests
import openpyxl
import requests
import sqlglot
import sqlglot.expressions as sqexp
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request, send_file
from google.cloud import bigquery as bq_lib
from rapidfuzz import fuzz

load_dotenv()


def _env(key: str) -> str:
    """Read a required variable set in .env — raises clearly if missing."""
    val = os.environ.get(key)
    if val is None:
        raise RuntimeError(f"Required config '{key}' is not set — add it to your .env file.")
    return val


GCP_PROJECT = _env("GCP_PROJECT")
BQ_LOCATION = _env("BQ_TRANSLATION_LOCATION")
BQ_TRANSLATE_URL = (
    f"https://bigquerymigration.googleapis.com/v2/"
    f"projects/{GCP_PROJECT}/locations/{BQ_LOCATION}:translateQuery"
)

GEMINI_LOCATION = _env("GEMINI_LOCATION")
GEMINI_MODEL    = _env("GEMINI_MODEL")
GEMINI_URL = (
    f"https://{GEMINI_LOCATION}-aiplatform.googleapis.com/v1/"
    f"projects/{GCP_PROJECT}/locations/{GEMINI_LOCATION}/"
    f"publishers/google/models/{GEMINI_MODEL}:generateContent"
)

MAX_WORKERS       = int(_env("MAX_WORKERS"))
_MAX_UPLOAD_BYTES = int(_env("MAX_UPLOAD_MB")) * 1024 * 1024

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = _MAX_UPLOAD_BYTES


def get_access_token() -> str:
    credentials, _ = google.auth.default(
        scopes=["https://www.googleapis.com/auth/cloud-platform"]
    )
    credentials.refresh(google.auth.transport.requests.Request())
    return credentials.token


def _call_gemini(prompt: str) -> str:
    """Call Gemini 2.5 Flash via Vertex AI and return the raw text response."""
    credentials, _ = google.auth.default(
        scopes=["https://www.googleapis.com/auth/cloud-platform"]
    )
    credentials.refresh(google.auth.transport.requests.Request())
    headers = {
        "Authorization": f"Bearer {credentials.token}",
        "Content-Type": "application/json",
    }
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {"responseMimeType": "application/json", "temperature": 0},
    }
    resp = requests.post(GEMINI_URL, headers=headers, json=payload, timeout=120)
    resp.raise_for_status()
    return resp.json()["candidates"][0]["content"]["parts"][0]["text"]


_DDL_SEPARATOR = "-- __QUERY_START__"  # survives translation; used to split DDL from SELECT in output


def _translate_one(token: str, sql: str, ddl: str = "") -> dict:
    """Call the BigQuery Migration translateQuery API for a single SQL string.

    If `ddl` is provided it is prepended so the translator resolves table
    references, and only the portion after the separator comment is returned.
    """
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    if ddl:
        query_input = ddl.strip() + "\n\n" + _DDL_SEPARATOR + "\n" + sql.strip()
    else:
        query_input = sql

    payload = {"sourceDialect": "ORACLE", "query": query_input}
    resp = requests.post(BQ_TRANSLATE_URL, headers=headers, json=payload, timeout=120)
    resp.raise_for_status()
    data = resp.json()

    records = data.get("translationResultDetails", {}).get("reportRecords", [])
    errors   = [r["message"] for r in records if r.get("severity") == "ERROR"]
    warnings = [r["message"] for r in records if r.get("severity") == "WARNING"]

    raw = data.get("translatedQuery", "")
    if ddl and _DDL_SEPARATOR in raw:
        translated = raw.split(_DDL_SEPARATOR, 1)[1].strip()
    else:
        translated = raw

    return {
        "translated_sql": translated,
        "errors": errors,
        "warnings": warnings,
    }


@app.route("/")
def index():
    return render_template("index.html", gcp_project=GCP_PROJECT)


@app.route("/api/list-bq-datasets", methods=["POST"])
def list_bq_datasets():
    body = request.get_json(silent=True) or {}
    project = body.get("project", GCP_PROJECT).strip()
    if not project:
        return jsonify({"error": "project is required"}), 400
    try:
        client = _get_bq_client(project)
        datasets = sorted(ds.dataset_id for ds in client.list_datasets())
        return jsonify({"datasets": datasets})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/translate-single", methods=["POST"])
def translate_single():
    body = request.get_json(silent=True) or {}
    sql = body.get("sql", "").strip()
    ddl = body.get("ddl", "").strip()
    if not sql:
        return jsonify({"error": "No SQL query provided"}), 400

    try:
        token = get_access_token()
    except Exception as exc:
        return jsonify({"error": f"Authentication failed: {exc}"}), 500

    try:
        result = _translate_one(token, sql, ddl=ddl)
    except requests.HTTPError as exc:
        return jsonify({"error": f"API error: {exc.response.text}"}), 502
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    if result["errors"]:
        status = "error"
    elif result["warnings"]:
        status = "warning"
    else:
        status = "ok"

    return jsonify({
        "translated_sql": result["translated_sql"],
        "errors": result["errors"],
        "warnings": result["warnings"],
        "status": status,
    })


@app.route("/api/preview-columns", methods=["POST"])
def preview_columns():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx / .xls files are supported"}), 400

    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    sheets = wb.sheetnames
    sheet_data = {}
    for name in sheets:
        ws = wb[name]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c) if c is not None else "" for c in row]
        sheet_data[name] = headers
    wb.close()
    return jsonify({"sheets": sheets, "columns": sheet_data})


@app.route("/api/translate", methods=["POST"])
def translate():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    sheet_name = request.form.get("sheet", "").strip()
    column_name = request.form.get("column", "").strip()
    start_row = max(1, int(request.form.get("start_row", 2)))
    ddl = request.form.get("ddl", "").strip()

    if not sheet_name or not column_name:
        return jsonify({"error": "sheet and column are required"}), 400

    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return jsonify({"error": f"Sheet '{sheet_name}' not found"}), 400
    ws = wb[sheet_name]

    # Find column index by header name
    col_index = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for i, h in enumerate(row):
            if str(h) == column_name:
                col_index = i
                break

    if col_index is None:
        return jsonify({"error": f"Column '{column_name}' not found"}), 400

    # Collect all rows
    rows_data = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=start_row, values_only=True), start=start_row
    ):
        val = row[col_index] if col_index < len(row) else None
        sql = str(val).strip() if val is not None else ""
        rows_data.append({"row": row_idx, "original_sql": sql})
    wb.close()

    non_empty = [(i, rd) for i, rd in enumerate(rows_data) if rd["original_sql"]]
    if not non_empty:
        return jsonify({"error": "No SQL queries found in the selected column"}), 400

    # Translate in parallel using ThreadPoolExecutor
    try:
        token = get_access_token()
    except Exception as exc:
        return jsonify({"error": f"Authentication failed: {exc}"}), 500

    translations: dict[int, dict] = {}
    try:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_idx = {
                executor.submit(_translate_one, token, rd["original_sql"], ddl): idx
                for idx, rd in non_empty
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    translations[idx] = future.result()
                except requests.HTTPError as exc:
                    translations[idx] = {
                        "translated_sql": "",
                        "errors": [f"API error: {exc.response.text}"],
                        "warnings": [],
                    }
                except Exception as exc:
                    translations[idx] = {
                        "translated_sql": "",
                        "errors": [str(exc)],
                        "warnings": [],
                    }
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    # Merge results back into original row order
    output_rows = []
    for i, rd in enumerate(rows_data):
        t = translations.get(i, {"translated_sql": "", "errors": [], "warnings": []})
        if not rd["original_sql"]:
            status = "empty"
        elif t["errors"]:
            status = "error"
        elif t["warnings"]:
            status = "warning"
        else:
            status = "ok"
        output_rows.append(
            {
                "row": rd["row"],
                "original_sql": rd["original_sql"],
                "translated_sql": t["translated_sql"],
                "errors": t["errors"],
                "warnings": t["warnings"],
                "status": status,
            }
        )

    return jsonify({"results": output_rows})


@app.route("/api/export", methods=["POST"])
def export():
    data = request.get_json()
    results = data.get("results", [])

    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translation Results"

    headers = [
        "Row", "Original Oracle SQL", "Translated BigQuery SQL",
        "Status", "Errors", "Warnings",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in results:
        ws.append(
            [
                r.get("row", ""),
                r.get("original_sql", ""),
                r.get("translated_sql", ""),
                r.get("status", ""),
                "; ".join(r.get("errors", [])),
                "; ".join(r.get("warnings", [])),
            ]
        )

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="bigquery_translated_queries.xlsx",
    )


# ── BigQuery schema helpers ────────────────────────────────────────────────────

# Keyword tokens that look like table names but aren't
_SQL_KEYWORDS = frozenset({
    "DUAL", "SELECT", "WHERE", "SET", "VALUES", "LATERAL", "UNNEST",
    "WITH", "RECURSIVE", "ON", "USING", "AS", "OUTER", "INNER",
    "LEFT", "RIGHT", "FULL", "CROSS", "NATURAL", "PARTITION",
})

def extract_oracle_table_refs(sql: str) -> list[str]:
    """Return a list of unique SCHEMA.TABLE (or TABLE) refs found in Oracle SQL."""
    s = re.sub(r"--[^\n]*", " ", sql)                          # strip line comments
    s = re.sub(r"/\*.*?\*/", " ", s, flags=re.DOTALL)          # block comments
    s = re.sub(r"'[^']*'", "''", s)                            # string literals
    pattern = (
        r"\b(?:FROM|JOIN|UPDATE|INTO|TABLE)\s+"
        r"([A-Za-z_]\w*(?:\.[A-Za-z_]\w*)?)"
    )
    refs, seen = [], set()
    for m in re.finditer(pattern, s, re.IGNORECASE):
        ref = m.group(1).upper()
        if ref not in _SQL_KEYWORDS and ref not in seen:
            seen.add(ref)
            refs.append(ref)
    return refs


# BigQuery → Oracle type mapping (no precision metadata in INFORMATION_SCHEMA)
_BQ_TO_ORACLE: dict[str, str] = {
    "STRING":      "VARCHAR2(4000)",
    "INT64":       "NUMBER(19)",
    "INTEGER":     "NUMBER(19)",
    "INT":         "NUMBER(19)",
    "SMALLINT":    "NUMBER(5)",
    "BIGINT":      "NUMBER(19)",
    "TINYINT":     "NUMBER(3)",
    "BYTEINT":     "NUMBER(3)",
    "FLOAT64":     "FLOAT",
    "FLOAT":       "FLOAT",
    "NUMERIC":     "NUMBER(38,9)",
    "DECIMAL":     "NUMBER(38,9)",
    "BIGNUMERIC":  "NUMBER(38,18)",
    "BIGDECIMAL":  "NUMBER(38,18)",
    "BOOL":        "NUMBER(1)",
    "BOOLEAN":     "NUMBER(1)",
    "DATE":        "DATE",
    "DATETIME":    "TIMESTAMP",
    "TIMESTAMP":   "TIMESTAMP WITH TIME ZONE",
    "TIME":        "VARCHAR2(8)",
    "BYTES":       "RAW(2000)",
    "JSON":        "CLOB",
    "GEOGRAPHY":   "CLOB",
    "INTERVAL":    "INTERVAL DAY(9) TO SECOND(6)",
    "ARRAY":       "CLOB",
    "STRUCT":      "CLOB",
    "RECORD":      "CLOB",
}

def bq_type_to_oracle(data_type: str) -> str:
    base = data_type.upper().split("<")[0].split("(")[0].strip()
    return _BQ_TO_ORACLE.get(base, f"VARCHAR2(4000) /* {data_type} */")


def _get_bq_client(project: str) -> bq_lib.Client:
    credentials, _ = google.auth.default(
        scopes=["https://www.googleapis.com/auth/cloud-platform"]
    )
    return bq_lib.Client(project=project, credentials=credentials)


@app.route("/api/bq-load-schema", methods=["POST"])
def bq_load_schema():
    """
    Load all tables + columns from INFORMATION_SCHEMA for a BQ dataset.
    Optionally accepts Oracle SQL to extract table references and find matches.
    Returns: all_tables, oracle_refs, matches, generated_ddl.
    """
    body = request.get_json(silent=True) or {}
    project = body.get("project", GCP_PROJECT).strip()
    dataset = body.get("dataset", "").strip()
    sql     = body.get("sql", "").strip()

    if not dataset:
        return jsonify({"error": "dataset is required"}), 400

    try:
        client = _get_bq_client(project)

        # 1. All tables ─ from INFORMATION_SCHEMA.TABLES
        tables_q = f"""
        SELECT table_name, table_type, ddl
        FROM `{project}.{dataset}.INFORMATION_SCHEMA.TABLES`
        WHERE table_type = 'BASE TABLE'
        ORDER BY table_name
        """
        table_rows = list(client.query(tables_q).result())

        if not table_rows:
            return jsonify({
                "all_tables": [], "oracle_refs": [],
                "matches": [], "generated_ddl": "",
            })

        table_names = [r.table_name for r in table_rows]
        table_ddls  = {r.table_name: r.ddl or "" for r in table_rows}

        # 2. All columns for those tables ─ from INFORMATION_SCHEMA.COLUMNS
        cols_q = f"""
        SELECT table_name, column_name, ordinal_position, data_type, is_nullable
        FROM `{project}.{dataset}.INFORMATION_SCHEMA.COLUMNS`
        WHERE table_name IN UNNEST(@tnames)
        ORDER BY table_name, ordinal_position
        """
        job_cfg = bq_lib.QueryJobConfig(
            query_parameters=[
                bq_lib.ArrayQueryParameter("tnames", "STRING", table_names)
            ]
        )
        col_rows = list(client.query(cols_q, job_config=job_cfg).result())

        # Group columns by table
        columns_by_table: dict[str, list[dict]] = {n: [] for n in table_names}
        for r in col_rows:
            columns_by_table[r.table_name].append({
                "column":   r.column_name,
                "position": r.ordinal_position,
                "bq_type":  r.data_type,
                "nullable": r.is_nullable == "YES",
                "oracle_type": bq_type_to_oracle(r.data_type),
            })

        all_tables = [
            {
                "name":    n,
                "columns": columns_by_table.get(n, []),
                "col_count": len(columns_by_table.get(n, [])),
                "bq_ddl":  table_ddls.get(n, ""),
            }
            for n in table_names
        ]

        # 3. Match Oracle table references to BQ table names
        oracle_refs = extract_oracle_table_refs(sql) if sql else []
        bq_upper = {n.upper(): n for n in table_names}  # case-insensitive lookup

        matches = []
        for ref in oracle_refs:
            table_part = ref.split(".")[-1].upper()
            bq_table = bq_upper.get(table_part)

            # Fallback: substring match (BQ table name contains Oracle table name)
            if not bq_table:
                for bq_name in table_names:
                    if table_part in bq_name.upper() or bq_name.upper() in table_part:
                        bq_table = bq_name
                        break

            matches.append({
                "oracle_ref": ref,
                "bq_table":   bq_table,
                "matched":    bq_table is not None,
                "columns":    columns_by_table.get(bq_table, []) if bq_table else [],
            })

        # 4. Generate Oracle DDL from matched tables
        ddl_parts = []
        for m in matches:
            if not m["matched"] or not m["columns"]:
                continue
            col_defs = []
            for col in m["columns"]:
                null_clause = "" if col["nullable"] else " NOT NULL"
                col_defs.append(f"  {col['column']} {col['oracle_type']}{null_clause}")
            ddl_parts.append(
                f"CREATE TABLE {m['oracle_ref']} (\n"
                + ",\n".join(col_defs)
                + "\n);"
            )

        return jsonify({
            "all_tables":    all_tables,
            "oracle_refs":   oracle_refs,
            "matches":       matches,
            "generated_ddl": "\n\n".join(ddl_parts),
        })

    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ── Step 1: Oracle SQL parser ──────────────────────────────────────────────────

_SKIP_TABLE_NAMES = frozenset({
    "DUAL", "SELECT", "WHERE", "SET", "VALUES", "LATERAL", "UNNEST",
    "WITH", "ON", "USING", "AS", "OUTER", "INNER", "LEFT", "RIGHT",
    "FULL", "CROSS", "NATURAL", "PARTITION",
})

_CLAUSE_TYPES = {
    "SELECT":   sqexp.Select,
    "WHERE":    sqexp.Where,
    "JOIN":     sqexp.Join,
    "ORDER BY": sqexp.Order,
    "GROUP BY": sqexp.Group,
    "HAVING":   sqexp.Having,
}


def parse_oracle_sql(sql: str) -> dict:
    """
    Parse Oracle SQL and return:
      tables  – [{schema, name, alias, full_ref}]
      columns – [{table_alias, name, contexts: [SELECT|WHERE|JOIN|...]}]
    """
    tables, columns = [], []
    parse_error = None

    try:
        ast = sqlglot.parse_one(
            sql, dialect="oracle",
            error_level=sqlglot.ErrorLevel.WARN,
        )

        # ── tables ────────────────────────────────────────────────────────
        seen = set()
        for t in ast.find_all(sqexp.Table):
            db_node = t.args.get("db")
            schema  = db_node.name.upper() if isinstance(db_node, sqexp.Identifier) else (str(db_node).upper() if db_node else "")
            name    = (t.name or "").upper()
            alias   = (t.alias or "").upper()
            if not name or name in _SKIP_TABLE_NAMES:
                continue
            full_ref = f"{schema}.{name}" if schema else name
            if full_ref not in seen:
                seen.add(full_ref)
                tables.append({"schema": schema, "name": name, "alias": alias, "full_ref": full_ref})

        # ── columns with clause context ───────────────────────────────────
        col_context: dict[tuple, set] = {}
        for label, cls in _CLAUSE_TYPES.items():
            for node in ast.find_all(cls):
                for col in node.find_all(sqexp.Column):
                    col_name  = (col.name  or "").upper()
                    tbl_alias = (col.table or "").upper()
                    if not col_name or col_name == "ROWNUM":
                        continue
                    key = (tbl_alias, col_name)
                    col_context.setdefault(key, set()).add(label)

        for (tbl_alias, col_name), contexts in col_context.items():
            columns.append({
                "table_alias": tbl_alias,
                "name":        col_name,
                "contexts":    sorted(contexts),
            })

    except Exception as exc:
        parse_error = str(exc)
        for ref in extract_oracle_table_refs(sql):
            parts = ref.split(".")
            tables.append({"schema": parts[0] if len(parts) > 1 else "", "name": parts[-1], "alias": "", "full_ref": ref})

    return {"tables": tables, "columns": columns, "parse_error": parse_error}


@app.route("/api/parse-oracle-sql", methods=["POST"])
def parse_oracle_sql_ep():
    body = request.get_json(silent=True) or {}
    sql  = body.get("sql", "").strip()
    if not sql:
        return jsonify({"error": "No SQL provided"}), 400
    return jsonify(parse_oracle_sql(sql))


# ── Step 3: Fuzzy matching ─────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Normalise identifier: lowercase, collapse underscores."""
    return re.sub(r"[_\s]+", "", s.lower())


def _score(a: str, b: str) -> int:
    return max(fuzz.ratio(a.upper(), b.upper()), fuzz.ratio(_norm(a), _norm(b)))


def match_tables(oracle_tables: list[dict], bq_tables: list[dict]) -> list[dict]:
    bq_by_name = {t["name"]: t for t in bq_tables}
    bq_names   = list(bq_by_name.keys())

    results = []
    for ot in oracle_tables:
        oracle_name = ot["name"]
        scored = sorted(
            [{"name": n, "score": _score(oracle_name, n)} for n in bq_names],
            key=lambda x: -x["score"],
        )
        best = scored[0] if scored else None
        bq_match = best["name"] if best and best["score"] >= 50 else None
        results.append({
            "oracle_ref":    ot["full_ref"],
            "oracle_schema": ot["schema"],
            "oracle_table":  oracle_name,
            "oracle_alias":  ot["alias"],
            "bq_table":      bq_match,
            "score":         best["score"] if best else 0,
            "matched":       bq_match is not None,
            "candidates":    scored[:6],
            "bq_columns":    bq_by_name[bq_match]["columns"] if bq_match else [],
        })
    return results


def match_columns(oracle_col_names: list[str], bq_columns: list[dict]) -> list[dict]:
    bq_names  = [c["column"] for c in bq_columns]
    bq_by_name = {c["column"]: c for c in bq_columns}

    results = []
    for oracle_col in oracle_col_names:
        scored = sorted(
            [{"name": n, "score": _score(oracle_col, n)} for n in bq_names],
            key=lambda x: -x["score"],
        )
        best     = scored[0] if scored else None
        bq_col   = best["name"] if best and best["score"] >= 40 else None
        col_info = bq_by_name.get(bq_col) if bq_col else {}
        results.append({
            "oracle_col":  oracle_col,
            "bq_col":      bq_col,
            "score":       best["score"] if best else 0,
            "bq_type":     col_info.get("bq_type", ""),
            "oracle_type": col_info.get("oracle_type", ""),
            "candidates":  scored[:6],
        })
    return results


def _match_schema_with_gemini(
    oracle_tables: list[dict],
    oracle_columns: list[dict],
    bq_tables: list[dict],
) -> dict:
    """Use Gemini 2.5 Flash to semantically match Oracle tables/columns to BigQuery."""
    bq_by_name  = {t["name"]: t for t in bq_tables}
    bq_upper    = {t["name"].upper(): t["name"] for t in bq_tables}
    bq_all_names = [t["name"] for t in bq_tables]
    oracle_by_ref = {t["full_ref"]: t for t in oracle_tables}

    # ── Table matching ──────────────────────────────────────────────────────
    table_prompt = (
        "You are a database schema mapping expert.\n"
        "Match each Oracle table reference to the most semantically similar BigQuery table.\n\n"
        f"Oracle tables: {json.dumps([t['full_ref'] for t in oracle_tables])}\n"
        f"BigQuery tables: {json.dumps(bq_all_names)}\n\n"
        "Return ONLY a JSON array. Each element must have:\n"
        '  "oracle_ref": string (exactly as provided)\n'
        '  "bq_table": string (exactly as provided) or null if no good match\n'
        '  "score": integer 0–100 (confidence)\n'
        '  "reasoning": string (one short sentence)\n'
        "Include every oracle_ref from the input."
    )

    table_matches: list[dict] = []
    try:
        gemini_tbls = json.loads(_call_gemini(table_prompt))
        seen_refs: set[str] = set()

        for item in gemini_tbls:
            ref = item.get("oracle_ref", "")
            seen_refs.add(ref)
            ot = oracle_by_ref.get(ref, {})

            # Validate BQ name exists (case-insensitive)
            bq_name = item.get("bq_table")
            if bq_name:
                bq_name = bq_upper.get(bq_name.upper())  # None if not found

            score   = max(0, min(100, int(item.get("score", 0))))
            bq_info = bq_by_name.get(bq_name) if bq_name else None

            table_matches.append({
                "oracle_ref":    ref or ot.get("full_ref", ""),
                "oracle_schema": ot.get("schema", ""),
                "oracle_table":  ot.get("name", ""),
                "oracle_alias":  ot.get("alias", ""),
                "bq_table":      bq_name,
                "score":         score,
                "matched":       bq_name is not None,
                "candidates":    [],
                "bq_columns":    bq_info["columns"] if bq_info else [],
                "reasoning":     item.get("reasoning", ""),
            })

        # Ensure Gemini didn't silently omit any Oracle tables
        for ot in oracle_tables:
            if ot["full_ref"] not in seen_refs:
                table_matches.append({
                    "oracle_ref": ot["full_ref"], "oracle_schema": ot["schema"],
                    "oracle_table": ot["name"],   "oracle_alias":  ot["alias"],
                    "bq_table": None, "score": 0, "matched": False,
                    "candidates": [], "bq_columns": [], "reasoning": "",
                })

    except Exception:
        # Graceful fallback to fuzzy matching
        table_matches = match_tables(oracle_tables, bq_tables)

    # ── Column matching (one Gemini call per matched table pair) ────────────
    col_matches_by_table: dict[str, list] = {}

    for tm in table_matches:
        if not tm["matched"] or not tm["bq_table"]:
            continue

        bq_tbl       = tm["bq_table"]
        bq_cols      = tm["bq_columns"]  # [{column, bq_type, oracle_type, nullable}]
        bq_col_upper = {c["column"].upper(): c["column"] for c in bq_cols}
        bq_col_info  = {c["column"]: c for c in bq_cols}

        # Gather Oracle columns relevant to this table
        alias_set = {tm["oracle_alias"], tm["oracle_table"], ""}
        oracle_col_names = sorted({
            c["name"] for c in oracle_columns
            if c["table_alias"] in alias_set
        })
        if not oracle_col_names and len(table_matches) == 1:
            oracle_col_names = sorted({c["name"] for c in oracle_columns})

        if not oracle_col_names:
            col_matches_by_table[bq_tbl] = []
            continue

        col_prompt = (
            "Match Oracle column names to BigQuery column names.\n"
            f"Oracle table: {tm['oracle_ref']}  →  BigQuery table: {bq_tbl}\n\n"
            f"Oracle columns: {json.dumps(oracle_col_names)}\n"
            f"BigQuery columns: {json.dumps([c['column'] for c in bq_cols])}\n\n"
            "Return ONLY a JSON array. Each element must have:\n"
            '  "oracle_col": string (exactly as provided)\n'
            '  "bq_col": string (exactly as provided) or null if no match\n'
            '  "score": integer 0–100 (confidence)\n'
            "Include every oracle_col from the input."
        )

        try:
            gemini_cols = json.loads(_call_gemini(col_prompt))
            seen_cols: set[str] = set()
            col_matches: list[dict] = []

            for item in gemini_cols:
                oracle_col = item.get("oracle_col", "")
                seen_cols.add(oracle_col)

                bq_col = item.get("bq_col")
                if bq_col:
                    bq_col = bq_col_upper.get(bq_col.upper())  # None if not found

                score = max(0, min(100, int(item.get("score", 0))))
                info  = bq_col_info.get(bq_col, {}) if bq_col else {}

                col_matches.append({
                    "oracle_col":  oracle_col,
                    "bq_col":      bq_col,
                    "score":       score,
                    "bq_type":     info.get("bq_type", ""),
                    "oracle_type": info.get("oracle_type", ""),
                    "candidates":  [],
                })

            # Fill in any oracle columns Gemini missed
            for oc in oracle_col_names:
                if oc not in seen_cols:
                    col_matches.append({
                        "oracle_col": oc, "bq_col": None, "score": 0,
                        "bq_type": "", "oracle_type": "", "candidates": [],
                    })

            col_matches_by_table[bq_tbl] = col_matches

        except Exception:
            # Graceful fallback to fuzzy matching for this table pair
            col_matches_by_table[bq_tbl] = match_columns(oracle_col_names, bq_cols)

    return {
        "table_matches":        table_matches,
        "col_matches_by_table": col_matches_by_table,
    }


@app.route("/api/match-schema", methods=["POST"])
def match_schema_ep():
    """
    Input: { oracle_tables, oracle_columns, bq_tables }
    Returns: { table_matches, col_matches_by_table }
    """
    body           = request.get_json(silent=True) or {}
    oracle_tables  = body.get("oracle_tables", [])
    oracle_columns = body.get("oracle_columns", [])
    bq_tables      = body.get("bq_tables", [])

    if not bq_tables:
        return jsonify({"error": "No BQ tables provided"}), 400

    try:
        result = _match_schema_with_gemini(oracle_tables, oracle_columns, bq_tables)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    return jsonify(result)


# ── Step 4: BigQuery query constructor ────────────────────────────────────────

def _rewrite_cols(sql: str, alias_to_bq: dict, col_map_by_bq: dict) -> str:
    """
    Rewrite Oracle SQL: substitute Oracle column names with matched BQ column names,
    and strip bare Oracle table-name qualifiers from TABLE.* expressions.
    alias_to_bq   : {"C": "CHDRPF", "E": "CLEXPF"}  (upper-cased)
    col_map_by_bq : {"CHDRPF": {"CHDRPFID": "policy_id"}, ...}  (upper-cased keys)
    """
    try:
        ast = sqlglot.parse_one(sql, dialect="oracle", error_level=sqlglot.ErrorLevel.WARN)
        for col in ast.find_all(sqexp.Column):
            col_name  = (col.name  or "").upper()
            tbl_alias = (col.table or "").upper()

            # TABLE.* → * when the qualifier is an Oracle table name (not a short alias)
            if col_name == "*" and tbl_alias in alias_to_bq:
                col.set("table", None)
                continue

            bq_tbl = alias_to_bq.get(tbl_alias)
            if not bq_tbl and not tbl_alias:
                for tbl, cmap in col_map_by_bq.items():
                    if col_name in cmap:
                        bq_tbl = tbl
                        break

            if bq_tbl:
                cmap = col_map_by_bq.get(bq_tbl, {})
                if col_name in cmap:
                    col.set("this", sqexp.to_identifier(cmap[col_name]))

        return ast.sql(dialect="oracle")
    except Exception:
        result = sql
        for cmap in col_map_by_bq.values():
            for o_col, bq_col in cmap.items():
                result = re.sub(
                    r"(?<![.\w])" + re.escape(o_col) + r"(?![.\w])",
                    bq_col, result, flags=re.IGNORECASE,
                )
        return result


@app.route("/api/construct-bq-query", methods=["POST"])
def construct_bq_query_ep():
    """
    Input:
      sql, project, dataset,
      table_mapping: [{oracle_ref, oracle_alias, bq_table, bq_columns:[{column,bq_type,oracle_type,nullable}]}]
      col_mapping:   {bq_table: [{oracle_col, bq_col}]}
    """
    body     = request.get_json(silent=True) or {}
    sql      = body.get("sql", "").strip()
    project  = body.get("project",  GCP_PROJECT).strip()
    dataset  = body.get("dataset",  "").strip()
    tbl_map  = body.get("table_mapping", [])
    col_map  = body.get("col_mapping",   {})

    if not sql:
        return jsonify({"error": "No SQL provided"}), 400

    try:
        token = get_access_token()
    except Exception as exc:
        return jsonify({"error": f"Auth failed: {exc}"}), 500

    # Build alias → BQ table (upper-cased)
    alias_to_bq: dict[str, str] = {}
    for tm in tbl_map:
        bq_tbl = (tm.get("bq_table") or "").upper()
        alias  = (tm.get("oracle_alias") or "").upper()
        tbl    = tm["oracle_ref"].split(".")[-1].upper()
        if alias:
            alias_to_bq[alias] = bq_tbl
        alias_to_bq[tbl] = bq_tbl

    # Flat col map: {BQ_TABLE_UPPER: {ORACLE_COL_UPPER: bq_col}}
    col_map_by_bq: dict[str, dict] = {
        bq_tbl.upper(): {
            m["oracle_col"].upper(): m["bq_col"]
            for m in mappings if m.get("bq_col")
        }
        for bq_tbl, mappings in col_map.items()
    }

    # Rewrite Oracle column names → matched BQ column names in the SQL
    rewritten = _rewrite_cols(sql, alias_to_bq, col_map_by_bq)

    # Generate Oracle DDL using BQ column names so the translator resolves the schema
    ddl_parts = []
    for tm in tbl_map:
        oracle_ref = tm["oracle_ref"]
        bq_cols    = tm.get("bq_columns", [])
        if bq_cols:
            col_defs = [
                f"  {c['column']} {c.get('oracle_type','VARCHAR2(4000)')}"
                + (" NOT NULL" if not c.get("nullable", True) else "")
                for c in bq_cols
            ]
            ddl_parts.append(f"CREATE TABLE {oracle_ref} (\n" + ",\n".join(col_defs) + "\n);")

    ddl = "\n\n".join(ddl_parts)

    # Translate with BigQuery API
    try:
        trans = _translate_one(token, rewritten, ddl=ddl)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    bq_sql = trans["translated_sql"]

    # Post-process: replace Oracle table refs with BQ fully-qualified names
    for tm in tbl_map:
        oracle_ref = tm["oracle_ref"]
        bq_tbl     = tm.get("bq_table")
        if not bq_tbl:
            continue
        bq_fqn = f"`{project}.{dataset}.{bq_tbl}`"

        # Strip "ORACLE_TABLE." qualifier from column refs first (TABLE.col → col, TABLE.* → *)
        # This handles cases where the BQ API or Oracle SQL uses the table name as column qualifier.
        for pat in [oracle_ref, oracle_ref.split(".")[-1]]:
            bq_sql = re.sub(
                r"(?<![`.\w])" + re.escape(pat) + r"\.",
                "", bq_sql, flags=re.IGNORECASE,
            )

        # Replace standalone table references with BQ FQN
        for pat in [oracle_ref, oracle_ref.split(".")[-1]]:
            bq_sql = re.sub(
                r"(?<![`.\w])" + re.escape(pat) + r"(?![`.\w])",
                bq_fqn, bq_sql, flags=re.IGNORECASE,
            )

    return jsonify({
        "bq_sql":        bq_sql,
        "rewritten_sql": rewritten,
        "ddl_used":      ddl,
        "errors":        trans["errors"],
        "warnings":      trans["warnings"],
        "status":        "error" if trans["errors"] else ("warning" if trans["warnings"] else "ok"),
    })


@app.route("/api/run-bq-query", methods=["POST"])
def run_bq_query():
    body = request.get_json(silent=True) or {}
    sql = body.get("sql", "").strip()
    project = body.get("project", GCP_PROJECT).strip()
    max_rows = min(int(body.get("max_rows", 500)), 5000)
    location = "us-central1"

    if not sql:
        return jsonify({"error": "No SQL provided"}), 400

    try:
        client = _get_bq_client(project)
        query_job = client.query(sql, location=location)
        results = query_job.result(max_results=max_rows)

        columns = [{"name": f.name, "type": f.field_type} for f in results.schema]
        rows = []
        for row in results:
            rows.append([str(v) if v is not None else None for v in row.values()])

        return jsonify({
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


if __name__ == "__main__":
    app.run(
        debug=_env("FLASK_DEBUG").lower() == "true",
        host=_env("FLASK_HOST"),
        port=int(_env("FLASK_PORT")),
    )
